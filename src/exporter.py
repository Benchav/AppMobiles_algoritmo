"""
exporter.py
-----------
Capa de presentación y exportación.

Responsabilidades:
  - Generar gráficos detallados (por área, año, carrera) con porcentajes.
  - Exportar conteo a CSV.
  - Exportar equipos a Excel con diseño profesional (colores, bordes, anchos).
"""

import os
import textwrap

import matplotlib
matplotlib.use('Agg')                         # sin GUI — compatible con todos los SO
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config import (
    AREA_ECONOMICAS, AREA_HUMANIDADES, AREA_INGENIERIA, AREA_SISTEMAS,
    OUTPUT_COUNTS_FILE, OUTPUT_DIR, OUTPUT_TEAMS_FILE, CLEANED_FILE,
    get_area_from_major,
)

# ── Paleta de colores para áreas ────────────────────────────────────────────
AREA_COLORS = {
    AREA_SISTEMAS:     '#2563EB',   # azul
    AREA_INGENIERIA:   '#16A34A',   # verde
    AREA_ECONOMICAS:   '#D97706',   # ámbar
    AREA_HUMANIDADES:  '#9333EA',   # violeta
    'Otra':            '#64748B',   # gris
}

TEAM_HEADER_FILL = '1E3A5F'        # azul marino profundo
TEAM_ROW_FILLS = {
    AREA_SISTEMAS:     'DBEAFE',    # azul claro
    AREA_INGENIERIA:   'DCFCE7',    # verde claro
    AREA_ECONOMICAS:   'FEF9C3',    # ámbar claro
    AREA_HUMANIDADES:  'F3E8FF',    # violeta claro
    'Otra':            'F1F5F9',
}


# ────────────────────────────────────────────────────────────────────────────
# GRÁFICOS
# ────────────────────────────────────────────────────────────────────────────

def _wrap(labels, width=20):
    return ['\n'.join(textwrap.wrap(str(l), width)) for l in labels]


def generate_charts(df: pd.DataFrame) -> list[str]:
    """Genera 3 gráficos: por Área, por Año y por Carrera. Retorna lista de rutas."""
    saved = []
    total = len(df)

    # ── 1. Pastel por Área ───────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(9, 7), facecolor='#0F172A')
    ax.set_facecolor('#0F172A')

    area_counts = df['Area'].value_counts()
    colors = [AREA_COLORS.get(a, '#64748B') for a in area_counts.index]

    wedges, texts, autotexts = ax.pie(
        area_counts.values,
        labels=None,
        autopct=lambda p: f'{p:.1f}%\n({int(round(p * total / 100))})',
        colors=colors,
        startangle=140,
        pctdistance=0.75,
        wedgeprops=dict(width=0.55, edgecolor='#0F172A', linewidth=2),
    )
    for at in autotexts:
        at.set(fontsize=10, fontweight='bold', color='white')

    legend_patches = [
        mpatches.Patch(color=AREA_COLORS.get(a, '#64748B'),
                       label=f'{a}  —  {n} est. ({n/total*100:.1f}%)')
        for a, n in area_counts.items()
    ]
    ax.legend(handles=legend_patches, loc='lower center',
              bbox_to_anchor=(0.5, -0.18), ncol=1,
              fontsize=9, framealpha=0, labelcolor='white')

    ax.set_title('Distribución de Estudiantes por Área',
                 fontsize=14, fontweight='bold', color='white', pad=18)
    fig.tight_layout()
    p1 = os.path.join(OUTPUT_DIR, 'grafico_areas.png')
    fig.savefig(p1, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    saved.append(p1)
    print(f"  [OK] Grafico 1 (Areas) exportado: {p1}")

    # ── 2. Barras agrupadas por Área y Año ───────────────────────────────────
    pivot = (
        df.groupby(['Area', 'Año_Num']).size()
          .reset_index(name='n')
          .pivot(index='Area', columns='Año_Num', values='n')
          .fillna(0)
    )
    años_disponibles = sorted(pivot.columns.astype(int).tolist())
    n_areas = len(pivot)
    n_años  = len(años_disponibles)
    x = np.arange(n_areas)
    width = 0.7 / max(n_años, 1)

    fig, ax = plt.subplots(figsize=(12, 7), facecolor='#0F172A')
    ax.set_facecolor('#1E293B')
    ax.tick_params(colors='white')
    # ax.spines is a dict-like mapping of spine name -> Spine instance
    # iterar sobre los spines y fijar el color (evita usar un slicing inválido)
    for spine in ax.spines.values():
        spine.set_color('#334155')

    YEAR_COLORS = {3: '#38BDF8', 4: '#818CF8', 5: '#34D399', 0: '#94A3B8'}

    for i, año in enumerate(años_disponibles):
        vals = pivot[año].values if año in pivot.columns else np.zeros(n_areas)
        bars = ax.bar(
            x + (i - n_años / 2 + 0.5) * width,
            vals, width * 0.9,
            color=YEAR_COLORS.get(año, '#CBD5E1'),
            label=f'{año}° año' if año > 0 else 'Otro/N.A.',
            zorder=3
        )
        for bar, v in zip(bars, vals):
            if v > 0:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height() + 0.2, f'{int(v)}',
                        ha='center', va='bottom', fontsize=8,
                        fontweight='bold', color='white')

    ax.set_xticks(x)
    ax.set_xticklabels(_wrap(pivot.index, 20), fontsize=9, color='white')
    ax.set_ylabel('Cantidad de Estudiantes', color='#94A3B8', fontsize=10)
    ax.set_title('Estudiantes por Área y Año Universitario',
                 fontsize=14, fontweight='bold', color='white', pad=14)
    ax.yaxis.label.set_color('#94A3B8')
    ax.tick_params(axis='y', colors='#94A3B8')
    ax.grid(axis='y', color='#334155', linestyle='--', alpha=0.5, zorder=0)
    ax.legend(fontsize=9, framealpha=0, labelcolor='white')
    fig.tight_layout()
    p2 = os.path.join(OUTPUT_DIR, 'grafico_area_año.png')
    fig.savefig(p2, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    saved.append(p2)
    print(f"  [OK] Grafico 2 (Area x Anno) exportado: {p2}")

    # ── 3. Barras horizontales por Carrera ───────────────────────────────────
    carrera_counts = (
        df.groupby(['Area', 'Carrera_Normalizada']).size()
          .reset_index(name='n')
          .sort_values('n', ascending=True)
    )
    bar_colors = [AREA_COLORS.get(a, '#64748B') for a in carrera_counts['Area']]
    labels_wrap = _wrap(carrera_counts['Carrera_Normalizada'], 36)

    fig, ax = plt.subplots(figsize=(12, max(6, len(carrera_counts) * 0.55 + 2)),
                           facecolor='#0F172A')
    ax.set_facecolor('#1E293B')
    for spine in ax.spines.values():
        spine.set_color('#334155')

    bars = ax.barh(labels_wrap, carrera_counts['n'], color=bar_colors,
                   edgecolor='#0F172A', linewidth=0.5, zorder=3)

    for bar, v in zip(bars, carrera_counts['n']):
        pct = v / total * 100
        ax.text(bar.get_width() + 0.15, bar.get_y() + bar.get_height() / 2,
                f' {int(v)}  ({pct:.1f}%)',
                va='center', fontsize=8.5, fontweight='bold', color='white')

    ax.set_xlabel('Cantidad de Estudiantes', color='#94A3B8', fontsize=10)
    ax.set_title('Estudiantes por Carrera y Área',
                 fontsize=14, fontweight='bold', color='white', pad=14)
    ax.tick_params(colors='white')
    ax.tick_params(axis='x', colors='#94A3B8')
    ax.grid(axis='x', color='#334155', linestyle='--', alpha=0.5, zorder=0)

    legend_patches = [
        mpatches.Patch(color=c, label=a)
        for a, c in AREA_COLORS.items()
        if a in carrera_counts['Area'].values
    ]
    ax.legend(handles=legend_patches, loc='lower right',
              fontsize=8.5, framealpha=0, labelcolor='white')

    fig.tight_layout()
    p3 = os.path.join(OUTPUT_DIR, 'grafico_carreras.png')
    fig.savefig(p3, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    saved.append(p3)
    print(f"  [OK] Grafico 3 (Carreras) exportado: {p3}")

    return saved


# ────────────────────────────────────────────────────────────────────────────
# CSV DE CONTEO
# ────────────────────────────────────────────────────────────────────────────

def export_counts(df: pd.DataFrame):
    """Exporta tabla de conteo detallada a CSV."""
    total = len(df)
    conteo = (
        df.groupby(['Area', 'Carrera_Normalizada', 'Año'])
          .size()
          .reset_index(name='Cantidad')
    )
    conteo['Porcentaje (%)'] = (conteo['Cantidad'] / total * 100).round(2)
    conteo.columns = ['Área', 'Carrera', 'Año', 'Cantidad', 'Porcentaje (%)']
    conteo = conteo.sort_values(['Área', 'Carrera', 'Año'])
    conteo.to_csv(OUTPUT_COUNTS_FILE, index=False, encoding='utf-8-sig')
    print(f"  [OK] Conteo CSV exportado: {OUTPUT_COUNTS_FILE}")

    # Generar y guardar gráficos
    generate_charts(df)


# ────────────────────────────────────────────────────────────────────────────
# EXCEL PROFESIONAL DE EQUIPOS
# ────────────────────────────────────────────────────────────────────────────

def _thin_border(sides='all'):
    thin = Side(border_style='thin', color='CBD5E1')
    medium = Side(border_style='medium', color='94A3B8')
    if sides == 'all':
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    if sides == 'outer':
        return Border(left=medium, right=medium, top=medium, bottom=medium)
    return Border()


def _header_style(ws, cell, text, fg='1E3A5F', color='FFFFFF', size=11, bold=True):
    cell.value = text
    cell.font = Font(name='Calibri', bold=bold, color=color, size=size)
    cell.fill = PatternFill('solid', fgColor=fg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _thin_border()


def export_teams(teams_df: pd.DataFrame):
    """Exporta equipos conformados a Excel con diseño profesional."""
    # Protección: si no hay datos, evitar crear un Excel vacío o con errores
    if teams_df is None or teams_df.empty:
        print("  [WARN] export_teams: No hay datos de equipos para exportar. Se omitirá la generación del Excel.")
        return
    wb = Workbook()

    # ── Hoja 1: Todos los equipos ────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Equipos Conformados'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    # Fila 1: Título
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'Concurso Multidisciplinario de Aplicaciones Moviles 2026  -  Equipos Conformados'
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='0F172A')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Fila 2: Encabezados
    headers = ['Equipo', '#', 'Nombre Completo', 'Carnet', 'Carrera', 'Área', 'Año']
    header_colors = ['1E3A5F', '1E3A5F', '1E3A5F', '1E3A5F', '1E3A5F', '1E3A5F', '1E3A5F']
    for col_idx, (h, fg) in enumerate(zip(headers, header_colors), start=1):
        c = ws.cell(row=2, column=col_idx)
        _header_style(ws, c, h, fg=fg)
    ws.row_dimensions[2].height = 22

    # Filas de datos agrupadas por equipo
    current_team = None
    team_start_row = 3
    fill_alternado = False   # para sombrear equipo sí / equipo no

    for row_idx, record in enumerate(teams_df.itertuples(index=False), start=3):
        equipo = record.Equipo
        area   = getattr(record, 'Area', 'Otra')

        # Detectar cambio de equipo → separador visual
        if equipo != current_team:
            if current_team is not None:
                # Línea divisoria entre equipos
                for col in range(1, 8):
                    ws.cell(row=row_idx - 1, column=col).border = Border(
                        left=_thin_border().left,
                        right=_thin_border().right,
                        top=_thin_border().top,
                        bottom=Side(border_style='medium', color='64748B')
                    )
            current_team = equipo
            fill_alternado = not fill_alternado

        area_fill = TEAM_ROW_FILLS.get(area, 'F8FAFC')
        bg = area_fill if not fill_alternado else _lighten(area_fill)

        row_data = [equipo, record._1, record.Nombre, record.Carnet,
                    record.Carrera, getattr(record, 'Area', ''), record.Año]

        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else '')
            c.font = Font(name='Calibri', size=10,
                          bold=(col_idx == 1),
                          color='1E3A5F' if col_idx == 1 else '0F172A')
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center' if col_idx in (1, 2, 4, 7) else 'left',
                                    vertical='center', wrap_text=True)
            c.border = _thin_border()

        ws.row_dimensions[row_idx].height = 18

    # Anchos de columnas
    col_widths = [16, 5, 38, 18, 38, 30, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Resumen por Equipo ───────────────────────────────────────────
    ws2 = wb.create_sheet(title='Resumen por Equipo')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:D1')
    t = ws2['A1']
    t.value = 'Resumen de Equipos Conformados'
    t.font = Font(bold=True, size=13, color='FFFFFF')
    t.fill = PatternFill('solid', fgColor='0F172A')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    res_headers = ['Equipo', 'Total Integrantes', 'Áreas Representadas', 'Carreras']
    for ci, h in enumerate(res_headers, start=1):
        c = ws2.cell(row=2, column=ci)
        _header_style(ws2, c, h)

    resumen = (
        teams_df.groupby('Equipo', as_index=False)
                .apply(
                    lambda g: pd.Series({
                        'Total Integrantes': len(g),
                        'Areas Representadas': ', '.join(sorted(g['Area'].dropna().unique())),
                        'Carreras': ', '.join(sorted(g['Carrera'].dropna().unique())),
                    }),
                    include_groups=False
                )
    )

    # Ordenar numéricamente por el número del equipo cuando el nombre contiene dígitos
    # Ej: 'Equipo 1', 'Equipo 2', ..., 'Equipo 10' -> order ascendente por 1..10
    try:
        if 'Equipo' in resumen.columns:
            nums = pd.to_numeric(resumen['Equipo'].astype(str)
                                  .str.extract(r"(\d+)", expand=False), errors='coerce')
            if nums.notna().any():
                resumen['__EquipoNum'] = nums.fillna(10**9).astype(int)
                resumen = resumen.sort_values('__EquipoNum').drop(columns='__EquipoNum').reset_index(drop=True)
    except Exception:
        # No fallar si algo inesperado ocurre; dejar el orden original
        pass

    alt = False
    for ri, row in enumerate(resumen.itertuples(index=False), start=3):
        alt = not alt
        bg = 'EFF6FF' if alt else 'FFFFFF'
        vals = [row.Equipo,
                getattr(row, 'Total Integrantes', ''),
                getattr(row, 'Areas Representadas', ''),
                getattr(row, 'Carreras', '')]
        for ci, val in enumerate(vals, start=1):
            c = ws2.cell(row=ri, column=ci, value=str(val) if val else '')
            c.font = Font(name='Calibri', size=10)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = _thin_border()
        ws2.row_dimensions[ri].height = 20

    for ci, w in enumerate([16, 16, 55, 70], start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Hoja 3: Estadísticas ─────────────────────────────────────────────────
    ws3 = wb.create_sheet(title='Estadísticas')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:E1')
    t3 = ws3['A1']
    t3.value = 'Estadísticas de Inscripción'
    t3.font = Font(bold=True, size=13, color='FFFFFF')
    t3.fill = PatternFill('solid', fgColor='0F172A')
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    stat_headers = ['Área', 'Carrera', 'Año', 'Cantidad', 'Porcentaje (%)']
    for ci, h in enumerate(stat_headers, start=1):
        c = ws3.cell(row=2, column=ci)
        _header_style(ws3, c, h)

    total_est = len(teams_df['Nombre'].unique()) if not teams_df.empty else 1
    alt = False
    row_pos = 3
    for area_val, grp in teams_df.groupby('Area'):
        area_bg = TEAM_ROW_FILLS.get(area_val, 'F8FAFC')
        for carrera_val, cgrp in grp.groupby('Carrera'):
            for año_val, agrp in cgrp.groupby('Año'):
                alt = not alt
                bg = area_bg
                n = len(agrp)
                pct = round(n / total_est * 100, 2)
                for ci, val in enumerate([area_val, carrera_val, año_val, n, pct], start=1):
                    c = ws3.cell(row=row_pos, column=ci, value=val)
                    c.font = Font(name='Calibri', size=10)
                    c.fill = PatternFill('solid', fgColor=bg)
                    c.alignment = Alignment(horizontal='center' if ci in (3, 4, 5) else 'left',
                                            vertical='center')
                    c.border = _thin_border()
                    if ci == 5:
                        c.number_format = '0.00"%"'
                ws3.row_dimensions[row_pos].height = 18
                row_pos += 1

    for ci, w in enumerate([28, 38, 10, 12, 16], start=1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(OUTPUT_TEAMS_FILE)
    print(f"  [OK] Excel profesional exportado: {OUTPUT_TEAMS_FILE}")

    # ── Hoja opcional: No asignados (inscritos válidos pero no presentes en equipos)
    try:
        if CLEANED_FILE and os.path.exists(CLEANED_FILE):
            try:
                df_clean = pd.read_csv(CLEANED_FILE, sep=';', encoding='utf-8')
            except UnicodeDecodeError:
                df_clean = pd.read_csv(CLEANED_FILE, sep=';', encoding='latin1')

            # localizar columnas relevantes
            carnet_col_clean = next((c for c in df_clean.columns if 'carnet' in str(c).lower()), None)
            nombre_col_clean = next((c for c in df_clean.columns if 'nombre' in str(c).lower()), None)

            # columnas en teams_df
            teams_carnet_col = next((c for c in teams_df.columns if 'carnet' in str(c).lower()), None)
            teams_nombre_col = next((c for c in teams_df.columns if 'nombre' in str(c).lower()), None)

            no_asig = None
            if teams_carnet_col and carnet_col_clean:
                assigned = set(teams_df[teams_carnet_col].astype(str).str.strip().str.lower())
                mask = ~df_clean[carnet_col_clean].astype(str).str.strip().str.lower().isin(assigned)
                no_asig = df_clean.loc[mask].reset_index(drop=True)
            elif teams_nombre_col and nombre_col_clean:
                assigned = set(teams_df[teams_nombre_col].astype(str).str.strip().str.lower())
                mask = ~df_clean[nombre_col_clean].astype(str).str.strip().str.lower().isin(assigned)
                no_asig = df_clean.loc[mask].reset_index(drop=True)

            if no_asig is not None:
                ws_no = wb.create_sheet(title='No asignados')
                ws_no.sheet_view.showGridLines = False

                # Título
                ws_no.merge_cells('A1:E1')
                t = ws_no['A1']
                t.value = 'Estudiantes No Asignados'
                t.font = Font(bold=True, size=13, color='FFFFFF')
                t.fill = PatternFill('solid', fgColor='0F172A')
                t.alignment = Alignment(horizontal='center', vertical='center')
                ws_no.row_dimensions[1].height = 28

                # Resumen breve: total y desglose por área
                try:
                    import pandas as _pd
                    def _area_of_row(r):
                        carrera_val = None
                        if 'Carrera' in no_asig.columns:
                            carrera_val = r['Carrera']
                        elif 'Carrera_Normalizada' in no_asig.columns:
                            carrera_val = r['Carrera_Normalizada']
                        if _pd.isna(carrera_val):
                            return 'Otra'
                        return get_area_from_major(carrera_val)

                    area_counts = {}
                    for _, r in no_asig.iterrows():
                        a = _area_of_row(r)
                        area_counts[a] = area_counts.get(a, 0) + 1

                    summary_parts = [f'Total no asignados: {len(no_asig)}']
                    if area_counts:
                        summary_parts.append('Por área: ' + ', '.join(f'{k}={v}' for k, v in area_counts.items()))
                    summary_text = '  —  '.join(summary_parts)
                except Exception:
                    summary_text = f'Total no asignados: {len(no_asig)}'

                ws_no.merge_cells('A2:E2')
                s = ws_no['A2']
                s.value = summary_text
                s.font = Font(name='Calibri', size=11, color='FFFFFF')
                s.fill = PatternFill('solid', fgColor='0F172A')
                s.alignment = Alignment(horizontal='center', vertical='center')
                ws_no.row_dimensions[2].height = 20

                # Columnas a mostrar (ordenadas)
                cols = []
                for c in (nombre_col_clean, carnet_col_clean, 'Carrera', 'Carrera_Normalizada', 'Año'):
                    if c and c in no_asig.columns and c not in cols:
                        cols.append(c)
                if not cols:
                    cols = list(no_asig.columns)

                # Encabezados (fila 3, tras el resumen)
                header_colors = ['1E3A5F'] * len(cols)
                for col_idx, (h, fg) in enumerate(zip(cols, header_colors), start=1):
                    c = ws_no.cell(row=3, column=col_idx)
                    _header_style(ws_no, c, h, fg=fg)
                ws_no.row_dimensions[3].height = 20

                # Filas con alternado y color por area si es posible (datos a partir de fila 4)
                fill_alternado = False
                # Iterar por filas tomando los valores en el mismo orden que 'cols'
                for row_idx, row in enumerate(no_asig[cols].itertuples(index=False, name=None), start=4):
                    fill_alternado = not fill_alternado
                    # obtener area a partir de la Carrera si es posible
                    # usamos la serie en la fila original para inferir la carrera
                    # (no_asig[cols] mantiene el orden de columnas)
                    # localizar columna de carrera en cols si existe
                    try:
                        carrera_idx = cols.index('Carrera') if 'Carrera' in cols else (cols.index('Carrera_Normalizada') if 'Carrera_Normalizada' in cols else None)
                    except ValueError:
                        carrera_idx = None

                    carrera_val = None
                    if carrera_idx is not None:
                        carrera_val = row[carrera_idx]

                    area_val = get_area_from_major(carrera_val) if carrera_val is not None else None
                    area_bg = TEAM_ROW_FILLS.get(area_val, 'F8FAFC') if area_val else 'FFFFFF'
                    bg = area_bg if not fill_alternado else _lighten(area_bg)

                    for col_idx, val in enumerate(row, start=1):
                        display = '' if pd.isna(val) else val
                        cell_value = str(display) if display != '' else ''
                        c = ws_no.cell(row=row_idx, column=col_idx, value=cell_value)
                        c.font = Font(name='Calibri', size=10)
                        c.fill = PatternFill('solid', fgColor=bg)
                        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        c.border = _thin_border()
                    ws_no.row_dimensions[row_idx].height = 18

                # Anchos de columnas razonables
                widths = [38, 18, 38, 12, 10]
                for i, w in enumerate(widths[:len(cols)], start=1):
                    ws_no.column_dimensions[get_column_letter(i)].width = w

                # Guardar para incluir la hoja nueva
                wb.save(OUTPUT_TEAMS_FILE)
                print(f"  [OK] Hoja 'No asignados' añadida al Excel: {OUTPUT_TEAMS_FILE}")
    except Exception as exc:
        print(f"  [WARN] No se pudo generar la hoja 'No asignados': {exc}")


def _lighten(hex_color: str, factor: float = 0.85) -> str:
    """Aclara un color hex mezclándolo con blanco."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'{r:02X}{g:02X}{b:02X}'
